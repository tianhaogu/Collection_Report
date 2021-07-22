#!/usr/bin/env python3
"""
Mew? Standard AMR collection report.
"""
import os
import re
import shutil
import sys
import json
import pycountry
import hashlib

from argparse import ArgumentParser
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import suppress
from datetime import date, datetime
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from statistics import median, StatisticsError
from sqlalchemy import func

from appen_general_utils.databases.collect import (
    File,
    Pin,
    Project,
    User,
    DynamicPrompt,
    StaticPrompt,
)
from appen_general_utils.databases.connect import (
    User as ConnectUser,
    UserAttribute
)
from appen_general_utils.argparse.types import json_load
from appen_general_utils.apis.ip_api import Client as IP_Client
from appen_general_utils.rclone import rclone_copy

from appen_general_utils.databases.stats import Stat, File as StatFile
from appen_stats_utils.validators import validate
from PIL import Image, ExifTags, UnidentifiedImageError
import piexif
import pyheif
import pprint

EXIF_HEADERS = [
    "Make",
    "Model",
    "Orientation",
    "Software",
    "DateTime",
    "YCbCrPositioning",
    "Compression",
    "XResolution",
    "YResolution",
    "ResolutionUnit",
    "ExposureTime",
    "FNumber",
    "ExposureProgram",
    "ExifVersion",
    "DateTimeOriginal",
    "DateTimeDigitized",
    "ComponentsConfiguration",
    "CompressedBitsPerPixel",
    "ExposureBiasValue",
    "MaxApertureValue",
    "MeteringMode",
    "Flash",
    "FocalLength",
    "FlashpixVersion",
    "ColorSpace",
    "PixelXDimension",
    "PixelYDimension",
    "FileSource",
    "InteroperabilityIndex",
    "InteroperabilityVersion",
    "GPSLatitude",
    "GPSLongitude",
]

SESSION_HEADERS = [
    "Directory Name",
    "Pin",
    "Total items",
    "Recorded items",
    "Skipped items",
    "Rejected items",
    "Duration",
    "Date",
    "Completed",
    "Abandoned",
    "Email",
    "Device IP",
    "Device ID",
    "Device Model",
    "Device OS",
    "Country",
    "Country Code",
    "Region",
    "Region Name",
]

EXTRA_HEADERS = ["Store_Unique_ID", "Lat", "Lng",
                 "Storefront_URL", "Store_front_EXIF", "Storefront_Category",
                 "Business_Name_EXIF", "Business_Name_URL", "Business_Name_Comment",
                 "Business_Status_EXIF", "Business_Status_URL", "Business_Status", "Business_Status_Comment",
                 "Dine_In_EXIF", "Dine_In_URL", "Dine_Status",
                 "Takeout_EXIF", "Takeout_URL", "Takeout_Status",
                 "Has_Delivery_EXIF", "Has_Delivery_URL", "Has_Delivery_Status",
                 "Has_Instore_Shopping_EXIF", "Has_Instore_Shopping_URL", "Has_Instore_Shopping_Status",
                 "Curbside_Pickup_EXIF", "Curbside_Pickup_URL", "Curbside_Pickup_Status"]

STAT_HEADERS = ["Session", "File", "Reason"]

def md5(path):
    with open(path, "rb") as handle:
        h = hashlib.md5()
        while chunk := handle.read(8192):
            h.update(chunk)
        return h.hexdigest()

def try_get_cached_sessions(
        report_path, session_headers=SESSION_HEADERS, stat_headers=STAT_HEADERS
):
    sessions = {}

    try:
        wb = load_workbook(report_path)
    except FileNotFoundError:
        return sessions

    ws_sessions = wb.active
    for i, row in enumerate(ws_sessions):
        if i == 0:
            sheet_headers = set([x.value for x in row if x])
            if sheet_headers != set(session_headers):
                print(
                    "WARNING: Current sheet headers do not match headers provided.",
                    file=sys.stderr,
                )

                bak_path = report_path + ".bak"
                print("Backing up old report to {}".format(bak_path), file=sys.stderr)
                shutil.copy(report_path, bak_path)
                return sessions

            continue

        row = {
            "Session": {head: cell.value for head, cell in zip(session_headers, row)},
            "Stats": [],
        }

        if all(x is None for x in row["Session"].values()):
            continue

        row["stats"] = []
        sessions[row["Session"]["Directory Name"]] = row

    if not sessions:
        return sessions

    ws_heads = []
    ws_stats = wb["Stats"]
    for i, row in enumerate(ws_stats):
        if i == 0:
            ws_heads = [cell.value for cell in row]
            continue

        row = {head: cell.value for head, cell in zip(ws_heads, row)}

        if all(x is None for x in row.values()):
            continue

        if row["Session"] in sessions:
            try:
                sessions[row["Session"]]["Stats"].append(row)
            except KeyError:
                pass

    return sessions


def process_files(session, schema, stat_headers, exclude_corpus_codes=set(), median_stats=False):
    item_stats, stat_rows = {}, []
    prompt_types = set(["recording", "video", "image"])
    file_counts = (
        File.query.with_entities(
            func.count(1),
            func.count(1).filter(File.attributes["skipped"].astext == "true"),
            func.count(1).filter(
                File.attributes["prompttype"].astext.in_(prompt_types),
                File.attributes["skipped"].astext == "false",
            ),
        )
            .filter(File.session_id == session.id)
            .one()
    )

    item_stats["total_items"] = file_counts[0]
    item_stats["skipped_items"] = file_counts[1]
    item_stats["recorded_items"] = file_counts[2]

    file_paths = File.query.with_entities(
        func.replace(File._path, "/ac-efs/", "/audio-efs/"),
        File.attributes["prompttype"].astext,
        File.attributes["corpuscode"]
    ).filter(  # noqa
        File.attributes["prompttype"].astext.in_(prompt_types),
        File.attributes["skipped"].astext == "false",
        File.session_id == session.id,
    )

    item_stats["rejected_items"] = 0
    for file_path, prompt_type, corpus_code in file_paths:
        if schema is not None and corpus_code not in exclude_corpus_codes:
            stat = (
                Stat.query.join(StatFile)
                    .filter(StatFile.path == file_path)
                    .order_by(Stat.created.desc())
                    .first()
            )
            if stat:
                vesults = validate(stat.json, schema)
                reasons = vesults.get_reasons()
                if reasons:
                    reasons_dict = {
                        "Session": session.name,
                        "File": os.path.basename(file_path),
                        "Reason": ",".join(reasons),
                    }
                if reasons or median_stats:
                    for head in stat_headers[3:]:
                        stat_value = None
                        if prompt_type == "recording":
                            stat_value = stat.json.get(head)

                        elif head.startswith("{}/".format(prompt_type)):
                            inner_json = stat.json.get(prompt_type, {})
                            stat_value = inner_json.get(head.split("/")[1])

                        elif head.startswith("audio/") and prompt_type == "video":
                            inner_json = stat.json.get("audio", {})
                            stat_value = inner_json.get(head.split("/")[1])

                        if stat_value is None:
                            continue

                        if reasons:
                            reasons_dict[head] = stat_value

                        if median_stats:
                            item_stats[head] = item_stats.get(head, []) + [stat_value]

                    if reasons:
                        item_stats["rejected_items"] += 1
                        stat_rows.append(reasons_dict)

            else:
                item_stats["missing_stats"] = item_stats.get("missing_stats", 0) + 1

    return item_stats, stat_rows


def parse_age(age):
    try:
        if type(age) is datetime:
            born = age.date()

        else:
            born = parse(age, ignoretz=True).date()

        today = date.today()
        age = (
                today.year
                - born.year
                - ((today.month, today.day) < (born.month, born.day))
        )
    except Exception:
        age = "Unknown data format: {}".format(
            age
        )

    return age


def get_exif_data(exif, exif_headers):
    exif_data = {}
    for header in exif_headers:
        # Start with Interop as it's a tiny class with two attributes
        key = getattr(piexif.InteropIFD, header, None)
        d = "Interop"
        if key is None:
            key = getattr(piexif.ImageIFD, header, None)
            d = "0th"
        if key is None:
            key = getattr(piexif.GPSIFD, header, None)
            d = "GPS"
        if key is None:
            key = getattr(piexif.ExifIFD, header, None)
            d = "Exif"
        if key is None:
            if header == "InteroperabilityVersion":
                key = 2  # No idea why this isn't in here
                d = "Interop"
            else:
                raise ValueError(f"{header} is not a valid Exif tag")

        # val returns the value of several different properties of the exif of the current photo
        val = exif[d].get(key)

        with suppress(AttributeError, UnicodeDecodeError):
            val = val.decode()  # Strings are returned as bytes, but int/none etc are possible

        if header in ("DateTimeOriginal", "DateTimeDigitized", "DateTime"):
            with suppress(Exception):
                val = int((parse(val).timestamp() * 1000) * 1000)  # Nanoseconds? Why?

        # exif_data stores all exif information of the current photo as a json object
        exif_data[header] = val

    return exif_data


def parse_lng_lat(x):
    # ((121, 1), (59, 1), (3926, 100))
    # ((33, 1), (47, 1), (37131958, 1000000))
    degrees, minutes, seconds = x
    degrees = degrees[0] / degrees[1]
    minutes = minutes[0] / minutes[1]
    seconds = seconds[0] / seconds[1]
    return (degrees + (minutes / 60) + (seconds / 3600))


def get_rows(  # noqa: C901
        session,
        args,
        stat_headers=STAT_HEADERS,
        session_headers=SESSION_HEADERS,
        cache=None,
        input_corpus_codes=set(),
):
    cache_hit = False
    stat_rows = []
    session_row = {}

    if session.name in cache:
        # NOTE(harry): If this is slow -- https://gist.github.com/hest/8798884
        total_items = File.query.filter_by(session=session).count()
        if (
                cache[session.name]["Session"]["Completed"]
                or cache[session.name]["Session"]["Abandoned"]
                or cache[session.name]["Session"]["Total items"] == total_items
        ):
            if (
                    (
                            not cache[session.name]["Session"]["Completed"]
                            and not cache[session.name]["Session"]["Abandoned"]
                    ) and (
                    session.completed or session.abandoned
            )
            ):
                pass
            else:
                for cached_stats_row in cache[session.name]["Stats"]:
                    stat_rows.append(
                        {head: cached_stats_row.get(head, None) for head in stat_headers}
                    )

                session_row = {
                    head: cache[session.name]["Session"].get(head, None)
                    for head in session_headers
                }

                cache_hit = True

    if not cache_hit:
        pin, email = (
            Pin.query.with_entities(Pin.pin, User.email)
                .filter(Pin.id == session.pin_id)
                .join(User)
                .one_or_none()
        )

        item_stats, stat_rows = process_files(
            session,
            args.schema,
            stat_headers,
            median_stats=args.median_stats,
            exclude_corpus_codes=args.exclude_corpus_codes,
        )

        session_duration = session.duration
        if session_duration is None:  # This should only happen to video and image sessions
            session_paths = (
                File
                    .query
                    .with_entities(
                    func.replace(File._path, "/ac-efs/", "/audio-efs/")
                )
                    .filter(
                    File.session == session,
                    File.attributes["prompttype"].astext == "video"
                )
            ).all()

            session_duration = 0.0
            for file_path in session_paths:
                with suppress(TypeError):
                    file_duration = (Stat.query
                                     .with_entities(
                        Stat.json["video"]["duration"]
                    )
                                     .join(StatFile)
                                     .filter(StatFile.path == file_path)
                                     ).one_or_none()[0]

                    if file_duration:
                        session_duration += (file_duration / 1000)

        session_row = {
            "Directory Name": session.name,
            "Pin": pin,
            "Total items": item_stats["total_items"],
            "Recorded items": item_stats["recorded_items"],
            "Skipped items": item_stats["skipped_items"],
            "Rejected items": item_stats["rejected_items"],
            "Duration": session_duration,
            "Date": session.created,
            "Completed": session.completed,
            "Abandoned": session.abandoned,
            "Email": email
        }

        for prompt in session.prompts:
            attributes = prompt.attributes
            if attributes is None:
                continue

            for prompt_attribute in args.prompt_attributes:
                session_row[prompt_attribute] = attributes.get(prompt_attribute)

            if all([session_row.get(x) for x in args.prompt_attributes]):
                break

        if args.median_stats:
            session_row["missing_stats"] = item_stats.get("missing_stats", 0)
            for head in STAT_HEADERS[3:]:
                check = []
                for x in item_stats.get(head, [0]):
                    if isinstance(x, (float, int)):
                        check.append(x)
                    elif x in ("NaN", "Infinity"):
                        continue
                    else:
                        print(x, "unrecognised output of stats for", head, file=sys.stderr)

                try:
                    session_row[head] = median(check)
                except StatisticsError:
                    session_row[head] = 0

        ip_meta = {}
        device_info = session.device_info
        if device_info:
            session_row.update({k: ",".join(v) for k, v in device_info.items()})
            ip_client = IP_Client()
            for ip in device_info["ips"]:
                if ip not in ip_meta:
                    ip_meta[ip] = ip_client.get_meta(ip)

            session_row["Device IP"] = ",".join(device_info["ips"])

            ip_meta_rows = [
                ("Country", "country"),
                ("Country Code", "countryCode"),
                ("Region", "region"),
                ("Region Name", "regionName")
            ]
            session_row.update({
                title:
                    ",".join(
                        [
                            ip_meta[ip].get(ip_meta_key, "N/A")
                            for ip in device_info["ips"]
                        ]
                    )
                for title, ip_meta_key in ip_meta_rows
            })

    if not cache_hit and args.demographics:
        user_id = args.demographics["pattern"].search(session_row["Pin"])
        if user_id:
            user_id = int(user_id.group(0))
            user = ConnectUser.query.get(user_id)
            if user:
                session_row["Connect User ID"] = user.id
                session_row["Country"] = user.country
                session_row["State"] = user.state
                session_row["City"] = user.city
                session_row["Email"] = user.email

                for header, id_ in args.demographics["attributes"].items():
                    user_attribute = (
                        UserAttribute.query
                            .filter_by(
                            user_id=user.id,
                            attribute_id=id_,
                        )
                            .one_or_none()
                    )
                    value = None
                    if user_attribute:
                        value = user_attribute.value
                    session_row[header] = value

                if session_row.get("Age (ia)"):
                    # Use in demogs in case of conflict with inputs
                    session_row["Age (ia)"] = parse_age(session_row["Age (ia)"])
                elif session_row.get("Age") and "Age (ia)" not in args.demographics["attributes"].keys():
                    session_row["Age"] = parse_age(session_row["Age"])

                if session_row.get("age_bracket"):
                    # Allow for special substitution
                    session_row["age_bracket"] = str(parse_age(session_row["age_bracket"]))

    if not cache_hit and input_corpus_codes:
        files = (File.query
                 .with_entities(
            File._path
        )
                 .filter(
            File.attributes["corpuscode"].astext
                .in_(input_corpus_codes),
                File.session == session
        )
                 .all()
                 )
        for f in files:  # for every file in one session (various json, jpg files within one Aztec_ sub-path, but only pay attention to json files here)
            fpath = f[
                0]  # one 'f' returns a table with records, but here one entity (column, that is the file path, the json file), so f[0] returns this json file
            try:
                with open(fpath) as handle:
                    data = json.load(handle)  # data is the json array in the json file
            except (json.decoder.JSONDecodeError, UnicodeDecodeError):
                continue

            with suppress(KeyError):
                for input_prompt in data:  # input_prompts are the json objects in the json array
                    if input_prompt["user_input"]:
                        session_row[input_prompt["name"]] = input_prompt[
                            "user_input"]  # unique id (from input 1), so on...

                        # session_row["Correct_Location"] = input_prompt["user_input"]
                        # session_row["EV_Station_Status"] = input_prompt["user_input"]
                        # session_row["Network_Provider"] = input_prompt["user_input"]
                        # session_row["Prominent_Branding"] = input_prompt["user_input"]
                        # session_row["One_Distince_Business"] = input_prompt["user_input"]
                        # session_row["Floor_Number"] = input_prompt["user_input"]
                        # session_row["Number_of_Pedestals_Plugs"] = input_prompt["user_input"]
        if session_row.get("First_Language"):  # STX specific
            with suppress(AttributeError):
                session_row["First_Language"] = pycountry.languages.get(
                    alpha_3=session_row["First_Language"]
                ).name

        if session_row.get("Primary_home_language"):  # STX specific
            with suppress(AttributeError):
                session_row["Primary_home_language"] = pycountry.languages.get(
                    alpha_3=session_row["Primary_home_language"]
                ).name

    if not cache_hit and args.script_categories:
        for script_category in args.script_categories:
            title = script_category["title"]
            script_num = int(session.pin.script.script_num)
            for script_num_rule, value in script_category["rules"].items():
                if isinstance(script_num_rule, str):
                    # Need to debug this, cannot reproduce
                    print(script_num_rule, value, "is a string", file=sys.stderr)
                    continue
                # So far rule may be a range or a list with a single int
                if script_num in script_num_rule:
                    session_row[title] = value
                    break

    # NOTE(harry): As not *all* cache hits result in no-act, the session row can still be modifed
    #     So we need to perform the args.substitutions on every iteration regardless.
    for k, v in args.substitutions.items():
        if k not in session_row:
            continue

        with suppress(KeyError, AttributeError):
            # On KeyError we don't need to substitute
            session_row[k] = v[str(session_row[k]).strip()]

    country = session_row.get("Country")
    if country and args.countries and not cache_hit:
        _country = pycountry.countries.get(alpha_3=country)
        if _country is None:
            _country = pycountry.countries.get(alpha_2=country)
        if _country is None:
            _country = pycountry.countries.get(name=country)

        if _country is not None:
            if args.countries == "alpha_3":
                session_row["Country"] = _country.alpha_3
            elif args.countries == "alpha_2":
                session_row["Country"] = _country.alpha_2
            elif args.countries == "full_name":
                session_row["Country"] = _country.name
    dupe_counter = 1
    for file in sorted(session.files, key=lambda x: x.created):
        print("check1:   ", session.path)
        print("check1:   ", file.path)
        print("check1:   ", file.prompt_type)
        session_row["pin"] = session.pin.pin
        if file.prompt_type == "image":  # we need to deal with the 3 images (EV station, User Interface and Plug Photo)
            prompt = file.prompt
            try:
                img = Image.open(file.path)  # by the path entity in the files table we open all 3 images
                exif = piexif.load(img.info["exif"])  # exif is like json-style appendix information embedded in the photo
                print("Normal")
            except UnidentifiedImageError:  # Usual errors: KeyError, UnidentifiedImageError (heic) files
                print("UnidentifiedImageError")
                heif = True
                img = pyheif.read_heif(file.path)
                for metadata in img.metadata:
                    if metadata["type"] == "Exif":
                        exif = piexif.load(metadata["data"][6:])
            except Exception as e:
                print("Exception")
                raise e  # DELETE ME
                break
            exif_data = get_exif_data(exif, EXIF_HEADERS)  # exif_data is also of json-style
            prompt_name = None
            print("check2:   ", prompt.corpus_code)
            # check where the key "Prompt" exists in the attributes, if not, replace manually
            #if (prompt.attributes.get("Prompt", "missing_prompt") == "missing_prompt"):
                # only hold for this project that has exactly 3 types of photos
            if prompt.corpus_code == "1image1":
                prompt_name = "ev_station"
            elif prompt.corpus_code == "1image2":
                prompt_name = "user_interface"
            elif prompt.corpus_code == "1image3":
                prompt_name = "plug"
            else:
                prompt_name = "missing_prompt"
                if "missing_prompt" not in SESSION_HEADERS:
                    SESSION_HEADERS.extend(["missing_prompt_photo_exif", "missing_prompt_photo_url", "missing_prompt_photo_lat", "missing_prompt_photo_lng"])
                    # if prompt_name in session_row:
                    #     dupe_counter += 1
                    #     prompt_name = f"{prompt_name}_{dupe_counter}"
            # if key "Prompt exists" in the attributes, just use it
            # else:
            #     prompt_name = prompt.attributes.get("Prompt", "missing_prompt").lower().replace(" ", "_")
            # if prompt_name in session_row:
            #     dupe_counter += 1
            #     prompt_name = f"{prompt_name}_{dupe_counter}"

            print("check3:   ", prompt_name)

            device_info = file.attributes.get("deviceinfo")
            lat = None
            lng = None
            url_key = "%s_%s" % (str(prompt_name), "url")
            lat_key = "%s_%s" % (str(prompt_name), "lat")
            lng_key = "%s_%s" % (str(prompt_name), "lng")
            with suppress(TypeError):
                lat = parse_lng_lat(exif_data["GPSLatitude"])
                lng = parse_lng_lat(exif_data["GPSLongitude"])

            if device_info is not None and not all([lat, lng]):
                with suppress(KeyError):
                    lng = device_info["location"]["longitude"]
                    lat = device_info["location"]["latitude"]

            file_md5 = md5(file.path)
            session_row[f"{prompt_name}_photo_exif"] = json.dumps(exif_data, ensure_ascii=False)
            session_row[f"{prompt_name}_photo_url"] = file_md5
            session_row[f"{prompt_name}_photo_lat"] = lat
            session_row[f"{prompt_name}_photo_lng"] = lng
            with suppress(Exception):
                img.close()
        print("End of this session:         ", dupe_counter)
        dupe_counter += 1
    return session_row, stat_rows


def main():  # noqa: C901
    parser = ArgumentParser(description=__doc__)
    parser.add_argument("project_id", type=int, help="Appen Collect project number")
    parser.add_argument("-s", "--schema", type=json_load, help="Schema json file")
    parser.add_argument(
        "-d", "--demographics", type=json_load, help="Demographics json file"
    )
    parser.add_argument(
        "-c", "--script_categories", type=json_load, help="JSON with rules based on script #"
    )
    parser.add_argument(
        "-i", "--inputs", action="store_true",
        help="Include data from input prompts as columns"
    )
    parser.add_argument(
        "-b", "--bluetooth", action="store_true", help="Expect bluetooth info in logs"
    )
    parser.add_argument(
        "-n", "--no_upload", action="store_true", help="Do not upload report"
    )
    parser.add_argument(
        "-r", "--report_name", type=str, help="over-ride automatic report name"
    )
    parser.add_argument(
        "-m", "--median_stats",
        action="store_true", help="Include median values of stats in schema"
    )
    parser.add_argument(
        "--from_scratch", action="store_true", help="Back up and re-run from scratch"
    )
    parser.add_argument(
        "--exclude_corpus_codes", type=json_load, default=[],
        help="JSON array of corpus codes to exclude from stat checks"
    )
    parser.add_argument(
        "--substitutions", type=json_load, default={},
        help="JSON with rules for basic substitutions"
    )
    parser.add_argument(
        "--countries", type=str,
        choices=("alpha_2", "alpha_3", "full_name"),
        help="Format all values in the 'Country' column"
    )
    parser.add_argument(
        "--prompt_attributes", type=str, nargs="+", default=[],
        help="prompt attribute keys"
    )
    args = parser.parse_args()

    args.exclude_corpus_codes = set(args.exclude_corpus_codes)

    if any([
        x is not None
        for x in
        (
                args.script_categories,
                args.bluetooth,
                args.inputs,
                args.median_stats
        )
    ]):
        global SESSION_HEADERS

    if args.schema:
        global STAT_HEADERS
        properties = args.schema["properties"]
        for property_ in sorted(properties.keys()):
            if property_ in ("video", "audio", "image"):  # there's got to be a better way
                for property__ in properties[property_]["properties"].keys():
                    STAT_HEADERS.append("/".join([property_, property__]))
                    if args.median_stats:
                        SESSION_HEADERS.append("/".join([property_, property__]))

            else:
                STAT_HEADERS.append(property_)
                if args.median_stats:
                    SESSION_HEADERS.append(property_)

        if args.median_stats:
            SESSION_HEADERS.append("missing_stats")

    if args.demographics is not None:
        SESSION_HEADERS.extend(["Connect User ID", "Country", "State", "City"])
        for header in sorted(args.demographics["attributes"].keys()):
            SESSION_HEADERS.append(header)
        args.demographics["pattern"] = re.compile(args.demographics["pattern"])

    if args.script_categories is not None:
        for script_category in args.script_categories:
            SESSION_HEADERS.append(script_category["title"])

            for k in list(script_category["rules"].keys()):
                v = script_category["rules"][k]
                if not isinstance(k, str):
                    continue

                if k.isdecimal():
                    script_category["rules"][(int(k),)] = v
                    del script_category["rules"][k]

                elif re.match(r"\d+-\d+", k):
                    minimum, maximum = (int(x) for x in k.split("-"))
                    maximum += 1
                    script_category["rules"][range(minimum, maximum)] = v
                    del script_category["rules"][k]

                else:
                    print("Unrecognised script category rule: {}: {}".format(k, v), file=sys.stderr)
                    del script_category["rules"][k]

    project = Project.query.get(args.project_id)
    if args.report_name is not None:
        report_name = args.report_name
        if not args.report_name.endswith(".xlsx"):
            report_name = "{}.xlsx".format(args.report_name)
    else:
        report_name = "{}_{}_{}_{}_collection_report.xlsx".format(
            project.number, project.name, project.description, project.lang_code
        )

    input_corpus_codes = set()
    if args.bluetooth:  # To change once attributes change
        SESSION_HEADERS.extend(["Bluetooth Name", "Bluetooth Type"])

    if args.inputs:
        inputs = (
            StaticPrompt.query
                .with_entities(
                StaticPrompt.corpus_code,
                StaticPrompt.inputs
            )
                .filter(
                StaticPrompt.project == project,
                StaticPrompt.prompt_type == "input"
            )
        ).all()
        if not inputs:
            inputs = (
                DynamicPrompt.query
                    .with_entities(
                    DynamicPrompt.corpus_code,
                    DynamicPrompt.inputs
                )
                    .filter(
                    DynamicPrompt.project == project,
                    DynamicPrompt.prompt_type == "input"
                )
            ).all()

        if inputs:
            seen_names = set()
            for corpus_code, i in inputs:
                input_corpus_codes.add(corpus_code)
                for input_prompt in i:
                    name = input_prompt["name"]
                    if name == "Please leave feedback, if any, relating to business status below":
                        seen_names.add("business_status_comment")
                    else:
                        seen_names.add(name)

        #pprint.pprint(inputs)
        #pprint.pprint(seen_names)
        try:
            SESSION_HEADERS.extend(list(sorted(seen_names)))  # Ensure same order for cache
        except UnboundLocalError:
            print("WARNING: You have included an --inputs argument, but the script found no input prompts",
                  file=sys.stderr)

    SESSION_HEADERS.extend(["ev_station_photo_exif", "ev_station_photo_url", "ev_station_photo_lat", "ev_station_photo_lng"])
    SESSION_HEADERS.extend(["user_interface_photo_exif", "user_interface_photo_url", "user_interface_photo_lat", "user_interface_photo_lng"])
    SESSION_HEADERS.extend(["plug_photo_exif", "plug_photo_url", "plug_photo_lat", "plug_photo_lng"])

    SESSION_HEADERS.extend(args.prompt_attributes)

    doc_path = os.path.join(project.docs_path, "TempReport")
    report_path = os.path.join(doc_path, report_name)
    cache = {}
    if not args.from_scratch:
        cache = try_get_cached_sessions(report_path)

    wb = Workbook(write_only=True)

    ws_sessions = wb.create_sheet("Sessions")
    ws_sessions.append(SESSION_HEADERS)

    ws_stats = wb.create_sheet("Stats")
    ws_stats.append(STAT_HEADERS)

    futures = []
    with ThreadPoolExecutor(max_workers=6) as executor:
        for session in project.sessions:
            futures.append(
                executor.submit(
                    get_rows,
                    session,
                    args,
                    cache=cache,
                    input_corpus_codes=input_corpus_codes,
                )
            )

        for future in as_completed(futures):
            session_row, stat_rows = future.result()
            vs = []
            for header in SESSION_HEADERS:
                v = session_row.get(header)
                if type(v) is str:
                    v2 = ""
                    for x in v:
                        if ord(x) >= 32:
                            v2 += x
                    v = v2
                vs.append(v)

            ws_sessions.append(vs)
            for stat_row in stat_rows:
                ws_stats.append([stat_row.get(header) for header in STAT_HEADERS])

    wb.save(report_path)

    if not args.no_upload:
        # We'll be moving to this new location gradually.. once all old projects end.
        dirname = "/Data Collection - AMR/{}".format(project.name)
        result = rclone_copy(report_path, drivename="report:", dirname=dirname)
        if result.returncode != 0:
            raise ValueError("ERROR {}: {}".format(result.returncode, result.err))


if __name__ == "__main__":
    main()
